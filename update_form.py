#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
商品订购表单自动更新脚本

功能：
1. 读取Excel文件中的商品数据
2. 提取商品图片
3. 价格自动+10
4. 生成products_with_images.json
5. 推送到GitHub

使用方法：
    python update_form.py [excel文件名]
    
示例：
    python update_form.py 20260510截止.xlsx
"""

import openpyxl
import json
import os
import sys
import shutil
import subprocess
from pathlib import Path


def extract_price(price_str):
    """从价格字符串中提取数字并+10"""
    if not price_str:
        return ""
    
    # 移除货币符号和空格
    price_str = str(price_str).replace('¥', '').replace('$', '').replace(',', '').strip()
    
    # 尝试提取数字
    try:
        # 如果包含范围（如 30-40），取平均值
        if '-' in price_str:
            parts = price_str.split('-')
            if len(parts) == 2:
                low = float(parts[0].strip())
                high = float(parts[1].strip())
                avg = (low + high) / 2
                return str(int(avg + 10))
        
        # 直接转换
        price = float(price_str)
        return str(int(price + 10))
    except:
        return price_str


def extract_images_from_excel(excel_path, output_dir, no_image_rows=None):
    """从Excel中按顺序提取图片，跳过无图行，返回有图片的行号集合和格式映射"""
    if no_image_rows is None:
        no_image_rows = set()

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    os.makedirs(output_dir, exist_ok=True)

    for f in os.listdir(output_dir):
        if f.endswith(('.png', '.jpg', '.jpeg')):
            os.remove(os.path.join(output_dir, f))

    raw_images = []
    if hasattr(ws, '_images'):
        for idx, img in enumerate(ws._images):
            try:
                image_data = img._data()
                image_format = (img.format if hasattr(img, 'format') else 'png').lower()
                if image_format == 'jpeg':
                    image_format = 'jpg'
                raw_images.append({
                    'data': image_data,
                    'format': image_format,
                })
            except Exception as e:
                print(f'图片{idx+1}提取失败: {e}')

    rows_needing_images = []
    for row in range(2, ws.max_row + 1):
        if row not in no_image_rows:
            rows_needing_images.append(row)

    image_rows = set()
    image_formats = {}
    for i, img_info in enumerate(raw_images):
        if i < len(rows_needing_images):
            target_row = rows_needing_images[i]
            fmt = img_info['format']
            img_path = os.path.join(output_dir, f'product_{target_row}.{fmt}')
            with open(img_path, 'wb') as f:
                f.write(img_info['data'])
            image_rows.add(target_row)
            image_formats[target_row] = fmt
            print(f'图片已保存: product_{target_row}.{fmt} (对应第{target_row}行)')

    print(f'已提取 {len(raw_images)} 张图片')
    print(f'有图片的行号: {sorted(image_rows)}')
    return image_rows, image_formats


def process_excel(excel_path):
    """处理Excel文件并生成JSON"""
    print(f'正在处理: {excel_path}')

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    script_dir = os.path.dirname(os.path.abspath(__file__))
    images_dir = os.path.join(script_dir, 'images')

    no_image_rows = set()
    for row in range(2, ws.max_row + 1):
        remark_val = str(ws.cell(row, 7).value or '').strip()
        if remark_val == '无图':
            no_image_rows.add(row)

    image_rows, image_formats = extract_images_from_excel(excel_path, images_dir, no_image_rows)

    products = []
    for row in range(2, ws.max_row + 1):
        has_image = row in image_rows
        raw_price = ws.cell(row, 5).value
        adjusted_price = extract_price(raw_price)

        remark_val = str(ws.cell(row, 7).value or '').strip()
        if remark_val == '无图':
            remark_val = ''

        fmt = image_formats.get(row, 'png')
        product = {
            'id': row - 1,
            '发售日': str(ws.cell(row, 1).value)[:10] if ws.cell(row, 1).value else '',
            '截止日': str(ws.cell(row, 2).value)[:10] if ws.cell(row, 2).value else '',
            '厂商': ws.cell(row, 3).value or '',
            '品名': ws.cell(row, 4).value or '',
            '价格': adjusted_price,
            '定金': ws.cell(row, 6).value or '',
            '备注': remark_val,
            '图片': f'images/product_{row}.{fmt}' if has_image else ''
        }
        products.append(product)

    json_path = os.path.join(script_dir, 'products_with_images.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)

    print(f'已生成商品数据: {len(products)} 条')
    print(f'价格已自动+10')

    return products


def push_to_github():
    """推送到GitHub"""
    print('\n正在推送到GitHub...')
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)
    
    try:
        # 添加所有更改
        subprocess.run(['git', 'add', '.'], check=True)
        
        # 提交
        subprocess.run(['git', 'commit', '-m', '更新商品数据和图片'], check=True)
        
        # 推送
        subprocess.run(['git', 'push', 'origin', 'main'], check=True)
        
        print('✓ 推送成功！')
        return True
    except subprocess.CalledProcessError as e:
        print(f'推送失败: {e}')
        print('请检查Git配置或手动推送')
        return False


def main():
    """主函数"""
    # 获取Excel文件路径
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        # 查找项目目录中的Excel文件
        script_dir = os.path.dirname(os.path.abspath(__file__))
        excel_files = [f for f in os.listdir(script_dir) if f.endswith('.xlsx')]
        
        if not excel_files:
            print('错误: 未找到Excel文件')
            print('请将Excel文件放入项目文件夹，或指定文件路径')
            print('用法: python update_form.py [excel文件名]')
            sys.exit(1)
        
        excel_path = os.path.join(script_dir, excel_files[0])
        print(f'自动找到Excel文件: {excel_files[0]}')
    
    # 检查文件是否存在
    if not os.path.exists(excel_path):
        print(f'错误: 文件不存在: {excel_path}')
        sys.exit(1)
    
    # 处理Excel
    products = process_excel(excel_path)
    
    # 推送到GitHub
    push_to_github()
    
    print('\n' + '='*50)
    print('更新完成！')
    print(f'商品数量: {len(products)}')
    print('价格已自动+10')
    print('='*50)


if __name__ == '__main__':
    main()
